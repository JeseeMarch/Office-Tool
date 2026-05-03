import os
import pandas as pd
from openpyxl import load_workbook

# 1️⃣ 设定 Excel 文件路径
folder_path = r"D:/excel"
file1 = os.path.join(folder_path, "excel1.xlsx")  # 学生名单
file2 = os.path.join(folder_path, "excel2.xlsx")  # 全年级信息

# 2️⃣ 先确保 excel2.xlsx 里的公式被替换为数值
wb = load_workbook(file2, data_only=True)  # 只读取计算后的数值
ws = wb.active  # 选择默认工作表
for row in ws.iter_rows():
    for cell in row:
        cell.value = cell.value  # 强制写回数值，去除公式
wb.save(file2)
wb.close()

# 3️⃣ 读取 Excel 数据
df1 = pd.read_excel(file1, engine="openpyxl")
df2 = pd.read_excel(file2, engine="openpyxl")  # 现在 df2 中不会包含公式

# 4️⃣ 按 "学号" 合并数据
merged_df = df1.merge(df2, on="学号", how="left")

# 5️⃣ 保存合并后的 Excel
output_file = os.path.join(folder_path, "merged_result.xlsx")
merged_df.to_excel(output_file, index=False, engine="openpyxl")

print(f"✅ 合并后的文件已保存到: {output_file}")
